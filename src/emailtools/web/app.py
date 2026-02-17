"""FastAPI web application for INtelliBOX."""

from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, Request, HTTPException, Query, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import desc, func, case, or_

from emailtools.database import get_session
from emailtools.models import Action, Assignment, Email, ProcessingLog, RosterMember
from emailtools.reporter.generator import generate_report_data, get_cached_program_news, get_cached_structured_program_news, generate_enhanced_report
from emailtools.config import settings
from emailtools.settings_service import SettingsService

# Create FastAPI app
app = FastAPI(
    title="INtelliBOX",
    description="AI-Powered Email Action Tracking System",
    version="1.0.0"
)

# Setup templates and static files
template_dir = Path(__file__).parent / "templates"
static_dir = Path(__file__).parent / "static"
templates = Jinja2Templates(directory=str(template_dir))
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

# Inject as get_program_name to avoid clashing with context vars named program_name
templates.env.globals['get_program_name'] = lambda: SettingsService.get_setting('program_name', '')


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    """Main dashboard showing summary statistics."""
    with get_session() as session:
        # Get statistics
        total_emails = session.query(Email).count()
        total_actions = session.query(Action).count()

        unassigned_actions = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None)
        ).count()

        # Priority breakdown for unassigned actions
        high_priority = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None),
            Action.priority == "high"
        ).count()

        medium_priority = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None),
            Action.priority == "medium"
        ).count()

        low_priority = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None),
            Action.priority == "low"
        ).count()

        # Assignment statistics
        assigned_actions = session.query(Action).join(Assignment).count()

        completed_actions = session.query(Action).join(Assignment).filter(
            Assignment.status == "completed"
        ).count()

        # Get overdue actions (assigned or unassigned, exclude completed)
        today = datetime.utcnow().date()
        overdue_count = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()
        overdue_actions = session.query(Action).outerjoin(Assignment).join(Email).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).order_by(Action.due_date).limit(5).all()

        # Get high priority unassigned actions
        high_priority_actions = session.query(Action).outerjoin(Assignment).join(Email).filter(
            Assignment.id.is_(None),
            Action.priority == "high"
        ).order_by(Action.due_date.asc().nullslast()).limit(5).all()

        # Get recent assignments (last 5, exclude completed)
        recent_assignments = session.query(Assignment, Action).join(
            Action
        ).filter(
            Assignment.status == "assigned"
        ).order_by(desc(Assignment.assigned_at)).limit(5).all()

        # Get recently completed (last 5)
        recent_completions = session.query(Assignment, Action).join(
            Action
        ).filter(
            Assignment.status == "completed"
        ).order_by(desc(Assignment.assigned_at)).limit(5).all()

        # Get cached structured program news
        program_news_data = get_cached_structured_program_news(session)

        # Get latest email date for "last updated" info
        latest_email = session.query(Email).order_by(desc(Email.received_date)).first()
        last_email_date = latest_email.received_date if latest_email else None

        # Calculate time since last email
        time_since_last_email = None
        if last_email_date:
            delta = datetime.utcnow() - last_email_date
            if delta.total_seconds() < 3600:
                time_since_last_email = f"{int(delta.total_seconds() / 60)}m ago"
            elif delta.total_seconds() < 86400:
                time_since_last_email = f"{int(delta.total_seconds() / 3600)}h ago"
            else:
                time_since_last_email = f"{int(delta.total_seconds() / 86400)}d ago"

        return templates.TemplateResponse("dashboard.html", {
            "request": request,
            "total_emails": total_emails,
            "total_actions": total_actions,
            "unassigned_actions": unassigned_actions,
            "assigned_actions": assigned_actions,
            "completed_actions": completed_actions,
            "high_priority": high_priority,
            "medium_priority": medium_priority,
            "low_priority": low_priority,
            "overdue_count": overdue_count,
            "overdue_actions": overdue_actions,
            "high_priority_actions": high_priority_actions,
            "recent_assignments": recent_assignments,
            "recent_completions": recent_completions,
            "program_news": program_news_data,
            "last_email_date": last_email_date,
            "time_since_last_email": time_since_last_email,
            "current_time": datetime.utcnow(),
        })


@app.get("/actions", response_class=HTMLResponse)
async def list_actions(
    request: Request,
    priority: Optional[str] = None,
    assigned: Optional[str] = None,
    assignee: Optional[str] = None,
    search: Optional[str] = None,
    overdue: Optional[str] = None,
    page: int = Query(1, ge=1)
):
    """List all actions with filtering."""
    with get_session() as session:
        # Get stats for header
        total_actions = session.query(Action).count()
        unassigned_actions = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None)
        ).count()
        high_priority = session.query(Action).outerjoin(Assignment).filter(
            Action.priority == "high",
            Assignment.id.is_(None)
        ).count()
        medium_priority = session.query(Action).outerjoin(Assignment).filter(
            Action.priority == "medium",
            Assignment.id.is_(None)
        ).count()
        low_priority = session.query(Action).outerjoin(Assignment).filter(
            Action.priority == "low",
            Assignment.id.is_(None)
        ).count()

        # Get overdue count
        today = datetime.utcnow().date()
        overdue_count = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()

        # Base query
        query = session.query(Action).outerjoin(Assignment).join(Email)

        # Apply filters
        if priority:
            query = query.filter(Action.priority == priority)

        if assigned == "true":
            query = query.filter(Assignment.id.isnot(None))
        elif assigned == "false":
            query = query.filter(Assignment.id.is_(None))

        if assignee:
            query = query.filter(Assignment.assigned_to == assignee)

        if overdue == "true":
            query = query.filter(Action.due_date < today)

        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    Action.title.ilike(search_term),
                    Action.description.ilike(search_term),
                    Email.subject.ilike(search_term)
                )
            )

        # Order by priority and due date
        priority_order = case(
            (Action.priority == "high", 1),
            (Action.priority == "medium", 2),
            (Action.priority == "low", 3),
            else_=4
        )
        query = query.order_by(priority_order, Action.due_date.asc().nullslast())

        # Pagination
        per_page = 50
        total_count = query.count()
        total_pages = (total_count + per_page - 1) // per_page

        actions = query.offset((page - 1) * per_page).limit(per_page).all()

        # Get list of assignees for filter dropdown
        assignees = session.query(Assignment.assigned_to).distinct().order_by(Assignment.assigned_to).all()
        assignee_list = [a[0] for a in assignees if a[0]]

        # Get roster members for Quick Assign dropdown
        roster = session.query(RosterMember).filter_by(active=True).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()

        return templates.TemplateResponse("actions.html", {
            "request": request,
            "actions": actions,
            "priority": priority,
            "assigned": assigned,
            "assignee": assignee,
            "search": search,
            "page": page,
            "total_pages": total_pages,
            "total_count": total_count,
            "total_actions": total_actions,
            "unassigned_actions": unassigned_actions,
            "high_priority": high_priority,
            "medium_priority": medium_priority,
            "low_priority": low_priority,
            "overdue_count": overdue_count,
            "overdue": overdue,
            "assignee_list": assignee_list,
            "roster": roster,
            "current_time": datetime.utcnow()
        })


@app.get("/actions/{action_id}", response_class=HTMLResponse)
async def view_action(request: Request, action_id: int):
    """View details of a specific action."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()

        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        # Get assignment if exists
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()

        # Get other actions from the same email
        related_actions = session.query(Action).filter(
            Action.email_id == action.email_id,
            Action.id != action_id
        ).all()

        # Get roster for assignment dropdown
        roster = session.query(RosterMember).filter_by(active=True).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()

        # Get recent assignees as fallback when no roster
        recent_assignees = session.query(Assignment.assigned_to).distinct().order_by(
            desc(Assignment.assigned_at)
        ).limit(10).all()
        recent_assignee_list = [a[0] for a in recent_assignees if a[0]]

        return templates.TemplateResponse("action_detail.html", {
            "request": request,
            "action": action,
            "assignment": assignment,
            "related_actions": related_actions,
            "roster": roster,
            "recent_assignees": recent_assignee_list,
            "current_time": datetime.utcnow()
        })


@app.get("/emails", response_class=HTMLResponse)
async def list_emails(
    request: Request,
    search: Optional[str] = None,
    processed: Optional[str] = None,
    days: Optional[int] = None,
    page: int = Query(1, ge=1)
):
    """List all emails."""
    with get_session() as session:
        # Shared banner stats (same as all pages)
        unassigned_actions = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None)
        ).count()
        high_priority = session.query(Action).outerjoin(Assignment).filter(
            Action.priority == "high",
            Assignment.id.is_(None)
        ).count()
        today = datetime.utcnow().date()
        overdue_count = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()
        latest_email = session.query(Email).order_by(desc(Email.received_date)).first()
        last_email_date = latest_email.received_date if latest_email else None
        time_since_last_email = None
        if last_email_date:
            delta = datetime.utcnow() - last_email_date
            if delta.total_seconds() < 3600:
                time_since_last_email = f"{int(delta.total_seconds() / 60)}m ago"
            elif delta.total_seconds() < 86400:
                time_since_last_email = f"{int(delta.total_seconds() / 3600)}h ago"
            else:
                time_since_last_email = f"{int(delta.total_seconds() / 86400)}d ago"

        # Get stats
        total_emails = session.query(Email).count()
        processed_emails = session.query(Email).filter(Email.processed == True).count()
        unprocessed_emails = total_emails - processed_emails

        # Count emails with high priority actions
        emails_with_high_priority = session.query(Email).join(Action).filter(
            Action.priority == "high"
        ).distinct().count()

        # Base query
        query = session.query(Email)

        # Apply filters
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    Email.subject.ilike(search_term),
                    Email.from_address.ilike(search_term),
                    Email.body_text.ilike(search_term)
                )
            )

        if processed == "true":
            query = query.filter(Email.processed == True)
        elif processed == "false":
            query = query.filter(Email.processed == False)

        # Date range filter
        if days:
            cutoff_date = datetime.utcnow() - timedelta(days=days)
            query = query.filter(Email.received_date >= cutoff_date)

        # Order by received date
        query = query.order_by(desc(Email.received_date))

        # Pagination
        per_page = 50
        total_count = query.count()
        total_pages = (total_count + per_page - 1) // per_page

        emails = query.offset((page - 1) * per_page).limit(per_page).all()

        return templates.TemplateResponse("emails.html", {
            "request": request,
            "emails": emails,
            "search": search,
            "processed": processed,
            "days": days,
            "page": page,
            "total_pages": total_pages,
            "total_count": total_count,
            "total_emails": total_emails,
            "processed_emails": processed_emails,
            "unprocessed_emails": unprocessed_emails,
            "emails_with_high_priority": emails_with_high_priority,
            "unassigned_actions": unassigned_actions,
            "high_priority": high_priority,
            "overdue_count": overdue_count,
            "time_since_last_email": time_since_last_email,
            "current_time": datetime.utcnow()
        })


@app.get("/emails/{email_id}", response_class=HTMLResponse)
async def view_email(request: Request, email_id: int):
    """View details of a specific email."""
    with get_session() as session:
        email = session.query(Email).filter_by(id=email_id).first()

        if not email:
            raise HTTPException(status_code=404, detail="Email not found")

        # Get associated actions
        actions = session.query(Action).filter_by(email_id=email_id).all()

        # Count actions by status
        assigned_count = sum(1 for action in actions if action.assignments)
        unassigned_count = len(actions) - assigned_count
        completed_count = sum(1 for action in actions if action.assignments and action.assignments[0].status == 'completed')

        # Count by priority
        high_priority_count = sum(1 for action in actions if action.priority == 'high')

        return templates.TemplateResponse("email_detail.html", {
            "request": request,
            "email": email,
            "actions": actions,
            "assigned_count": assigned_count,
            "unassigned_count": unassigned_count,
            "completed_count": completed_count,
            "high_priority_count": high_priority_count,
            "current_time": datetime.utcnow()
        })


@app.get("/report", response_class=HTMLResponse)
async def view_report(request: Request, refresh: bool = False):
    """View AI-powered insights dashboard."""
    with get_session() as session:
        # Use enhanced report with caching
        report_data = generate_enhanced_report(session, days=7, force_refresh=refresh)

        # Calculate cache age in minutes for display
        if report_data.get("is_cached"):
            cache_age_minutes = int((datetime.utcnow() - report_data["generated_at"]).total_seconds() / 60)
        else:
            cache_age_minutes = 0

        return templates.TemplateResponse("report.html", {
            "request": request,
            "report": report_data,
            "cache_age_minutes": cache_age_minutes,
            "datetime": datetime  # Make datetime available in template
        })


@app.get("/api/stats")
async def get_stats():
    """API endpoint for statistics (for auto-refresh)."""
    with get_session() as session:
        return {
            "total_emails": session.query(Email).count(),
            "total_actions": session.query(Action).count(),
            "unassigned_actions": session.query(Action).outerjoin(Assignment).filter(
                Assignment.id.is_(None)
            ).count(),
            "high_priority": session.query(Action).outerjoin(Assignment).filter(
                Assignment.id.is_(None),
                Action.priority == "high"
            ).count()
        }


@app.post("/actions/{action_id}/assign")
async def assign_action(
    action_id: int,
    assigned_to: str = Form(...),
    notes: str = Form(""),
):
    """Assign an action to someone."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        # Check if already assigned
        existing = session.query(Assignment).filter_by(action_id=action_id).first()

        if existing:
            # Update existing assignment
            existing.assigned_to = assigned_to
            existing.notes = notes
            existing.assigned_at = datetime.utcnow()
        else:
            # Create new assignment
            assignment = Assignment(
                action_id=action_id,
                assigned_to=assigned_to,
                notes=notes,
                status="assigned"
            )
            session.add(assignment)

        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/priority")
async def update_priority(
    action_id: int,
    priority: str = Form(...),
):
    """Update action priority."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        if priority not in ["high", "medium", "low"]:
            raise HTTPException(status_code=400, detail="Invalid priority")

        action.priority = priority
        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/due-date")
async def update_due_date(
    action_id: int,
    due_date: str = Form(""),
):
    """Update action due date."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        # Parse date or clear if empty
        if due_date:
            from datetime import datetime as dt
            try:
                action.due_date = dt.strptime(due_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format")
        else:
            action.due_date = None

        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/complete")
async def complete_action(action_id: int):
    """Mark an action as completed."""
    with get_session() as session:
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()

        if not assignment:
            # Create assignment with completed status
            assignment = Assignment(
                action_id=action_id,
                assigned_to="Unknown",
                status="completed"
            )
            session.add(assignment)
        else:
            assignment.status = "completed"

        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/unassign")
async def unassign_action(action_id: int):
    """Remove assignment from an action."""
    with get_session() as session:
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()

        if assignment:
            session.delete(assignment)
            session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/status")
async def update_assignment_status(
    action_id: int,
    status: str = Form(...)
):
    """Update assignment status (assigned, completed)."""
    with get_session() as session:
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()

        if not assignment:
            raise HTTPException(status_code=404, detail="Action not assigned")

        if status not in ["assigned", "completed"]:
            raise HTTPException(status_code=400, detail="Invalid status")

        assignment.status = status
        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/edit")
async def edit_action(
    action_id: int,
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form("medium"),
    due_date: str = Form(""),
):
    """Update title, description, priority, and due date in one submit."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        action.title = title.strip()
        action.description = description.strip() if description.strip() else None

        if priority in ["high", "medium", "low"]:
            action.priority = priority

        if due_date and due_date.strip():
            from datetime import datetime as dt
            try:
                action.due_date = dt.strptime(due_date.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass
        else:
            action.due_date = None

        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/title")
async def update_action_title(
    action_id: int,
    title: str = Form(...)
):
    """Update action title."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        action.title = title
        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/description")
async def update_action_description(
    action_id: int,
    description: str = Form(...)
):
    """Update action description."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        action.description = description if description.strip() else None
        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/delete")
async def delete_action(action_id: int):
    """Delete an action."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        email_id = action.email_id
        session.delete(action)
        session.commit()

    return RedirectResponse(url=f"/emails/{email_id}", status_code=303)


@app.get("/emails/{email_id}/actions/new", response_class=HTMLResponse)
async def new_action_form(request: Request, email_id: int):
    """Show form to create a new action for an email."""
    with get_session() as session:
        email = session.query(Email).filter_by(id=email_id).first()
        if not email:
            raise HTTPException(status_code=404, detail="Email not found")

        return templates.TemplateResponse("action_new.html", {
            "request": request,
            "email": email
        })


@app.post("/emails/{email_id}/actions/new")
async def create_action(
    email_id: int,
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form("medium"),
    due_date: Optional[str] = Form(None),
    category: str = Form("")
):
    """Create a new action for an email."""
    with get_session() as session:
        email = session.query(Email).filter_by(id=email_id).first()
        if not email:
            raise HTTPException(status_code=404, detail="Email not found")

        # Parse due date if provided
        parsed_due_date = None
        if due_date and due_date.strip():
            try:
                from datetime import datetime as dt
                parsed_due_date = dt.strptime(due_date, '%Y-%m-%d')
            except ValueError:
                pass

        # Create new action
        action = Action(
            email_id=email_id,
            title=title,
            description=description if description.strip() else None,
            priority=priority,
            due_date=parsed_due_date,
            category=category if category.strip() else None,
            confidence_score=1.0  # Manual actions are 100% confident
        )

        session.add(action)
        session.commit()
        session.refresh(action)

    return RedirectResponse(url=f"/actions/{action.id}", status_code=303)


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, success: bool = False):
    """Settings page for configuring priority rules."""
    priority_config = SettingsService.get_priority_config()
    ai_config = SettingsService.get_ai_config()
    high_senders = priority_config.get('high_senders', [])
    high_keywords = priority_config.get('high_keywords', [])
    categories = ai_config.get('categories', SettingsService.DEFAULT_CATEGORIES)

    # Format categories as "name: description" lines for the textarea
    categories_text = '\n'.join(f"{c['name']}: {c['description']}" for c in categories)

    with get_session() as session:
        roster = session.query(RosterMember).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()

        return templates.TemplateResponse(
            "settings.html",
            {
                "request": request,
                "priority_default": priority_config.get('default_priority', 'medium'),
                "priority_days_threshold": priority_config.get('days_threshold', 5),
                "priority_high_senders_text": '\n'.join(high_senders),
                "priority_high_keywords_text": '\n'.join(high_keywords),
                "confidence_threshold": ai_config.get('confidence_threshold', 0.5),
                "categories_text": categories_text,
                "timezone": SettingsService.get_timezone(),
                "program_name": SettingsService.get_setting('program_name', ''),
                "success": success,
                "roster": roster,
            }
        )


@app.post("/settings")
async def save_settings(
    request: Request,
    priority_default: str = Form(...),
    priority_days_threshold: int = Form(...),
    priority_high_senders: str = Form(""),
    priority_high_keywords: str = Form(""),
    confidence_threshold: float = Form(0.5),
    ai_categories: str = Form(""),
    timezone: str = Form("America/Chicago"),
    program_name: str = Form("")
):
    """Save priority and AI settings."""
    # Parse textarea inputs (newline-separated) into lists
    senders = [s.strip() for s in priority_high_senders.split('\n') if s.strip()]
    keywords = [k.strip() for k in priority_high_keywords.split('\n') if k.strip()]

    # Parse categories using shared service method
    categories = SettingsService.parse_categories_text(ai_categories)

    # Save settings
    SettingsService.set_setting('priority_default', priority_default)
    SettingsService.set_setting('priority_days_threshold', priority_days_threshold)
    SettingsService.set_setting('priority_high_senders', senders)
    SettingsService.set_setting('priority_high_keywords', keywords)
    SettingsService.set_setting('ai_confidence_threshold', round(float(confidence_threshold), 2))
    SettingsService.set_setting('ai_categories', categories)
    SettingsService.set_setting('timezone', timezone)
    SettingsService.set_setting('program_name', program_name.strip())

    # Redirect with success flag
    return RedirectResponse(url="/settings?success=true", status_code=303)


@app.get("/roster", response_class=HTMLResponse)
async def view_roster(request: Request):
    """View and manage the program roster."""
    with get_session() as session:
        members = session.query(RosterMember).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()
        unassigned_actions = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None)
        ).count()
        high_priority = session.query(Action).outerjoin(Assignment).filter(
            Action.priority == "high", Assignment.id.is_(None)
        ).count()
        today = datetime.utcnow().date()
        overdue_count = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()
        return templates.TemplateResponse("roster.html", {
            "request": request,
            "members": members,
            "unassigned_actions": unassigned_actions,
            "high_priority": high_priority,
            "overdue_count": overdue_count,
            "current_time": datetime.utcnow(),
        })


@app.post("/roster/add")
async def add_roster_member(
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    force: str = Form(""),
):
    """Manually add a single roster member with duplicate and fuzzy-name checking."""
    import difflib

    first_name = first_name.strip()
    last_name = last_name.strip()
    email = email.strip().lower()
    full_name = f"{first_name} {last_name}".lower()

    with get_session() as session:
        # Exact email duplicate
        if session.query(RosterMember).filter_by(email=email).first():
            return RedirectResponse(
                f"/settings?roster_error={email}+is+already+in+the+roster#roster",
                status_code=303
            )

        # Fuzzy name check (skip if user confirmed with force=true)
        if not force:
            existing_names = [
                (m.id, f"{m.first_name} {m.last_name}")
                for m in session.query(RosterMember).all()
            ]
            for mid, name in existing_names:
                ratio = difflib.SequenceMatcher(None, full_name, name.lower()).ratio()
                if ratio >= 0.8:
                    return RedirectResponse(
                        f"/settings?roster_fuzzy={name}&roster_fn={first_name}"
                        f"&roster_ln={last_name}&roster_em={email}#roster",
                        status_code=303
                    )

        session.add(RosterMember(first_name=first_name, last_name=last_name, email=email))
        session.commit()

    return RedirectResponse("/settings?roster_added=1&roster_skipped=0", status_code=303)


@app.post("/roster/upload", response_class=HTMLResponse)
async def upload_roster(request: Request, file: UploadFile = File(...)):
    """Upload an Excel file to populate the roster."""
    import io
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        return RedirectResponse("/settings?roster_error=Please+upload+an+Excel+file+(.xlsx)", status_code=303)

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Find header row — look for first_name/last_name/email columns (case-insensitive)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        val = str(cell.value or "").strip().lower().replace(" ", "_")
        if val in ("first_name", "firstname", "first"):
            headers["first_name"] = col_idx
        elif val in ("last_name", "lastname", "last"):
            headers["last_name"] = col_idx
        elif val in ("email", "email_address", "work_email"):
            headers["email"] = col_idx

    if not all(k in headers for k in ("first_name", "last_name", "email")):
        return RedirectResponse(
            "/settings?roster_error=Excel+must+have+columns:+first_name,+last_name,+email",
            status_code=303
        )

    added = 0
    skipped = 0
    with get_session() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            first = str(row[headers["first_name"] - 1] or "").strip()
            last = str(row[headers["last_name"] - 1] or "").strip()
            email = str(row[headers["email"] - 1] or "").strip().lower()

            if not first or not last or not email or "@" not in email:
                skipped += 1
                continue

            existing = session.query(RosterMember).filter_by(email=email).first()
            if existing:
                existing.first_name = first
                existing.last_name = last
                existing.active = True
                skipped += 1
            else:
                session.add(RosterMember(first_name=first, last_name=last, email=email))
                added += 1
        session.commit()

    return RedirectResponse(f"/settings?roster_added={added}&roster_skipped={skipped}", status_code=303)


@app.post("/roster/{member_id}/delete")
async def delete_roster_member(member_id: int):
    """Remove a member from the roster."""
    with get_session() as session:
        member = session.query(RosterMember).filter_by(id=member_id).first()
        if member:
            name = member.full_name
            session.delete(member)
            session.commit()
            return RedirectResponse(f"/settings?roster_deleted={name}#roster", status_code=303)
    return RedirectResponse("/settings#roster", status_code=303)


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
