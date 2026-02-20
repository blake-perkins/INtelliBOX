"""Roster management routes."""

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse

from intellibox.models import RosterMember
from intellibox.utils.datetime_utils import utcnow
from intellibox.web.deps import get_session, templates
from intellibox.web.queries import get_banner_stats

router = APIRouter()


@router.get("/roster", response_class=HTMLResponse)
async def view_roster(request: Request):
    """View and manage the program roster."""
    with get_session() as session:
        members = session.query(RosterMember).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()
        stats = get_banner_stats(session)

        return templates.TemplateResponse("roster.html", {
            "request": request,
            "members": members,
            "unassigned_actions": stats["unassigned_actions"],
            "high_priority": stats["high_priority"],
            "overdue_count": stats["overdue_count"],
            "current_time": utcnow(),
        })


@router.post("/roster/add")
async def add_roster_member(
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    force: str = Form(""),
):
    """Manually add a single roster member with duplicate and fuzzy-name checking."""
    import difflib

    first_name = first_name.strip()
    last_name = last_name.strip()
    email = email.strip().lower()

    if not email:
        return RedirectResponse(
            "/settings?roster_error=Email+is+required#roster",
            status_code=303
        )

    full_name = f"{first_name} {last_name}".lower()

    with get_session() as session:
        # Exact email duplicate
        if session.query(RosterMember).filter_by(email=email).first():
            return RedirectResponse(
                f"/settings?roster_error={email}+is+already+in+the+roster#roster",
                status_code=303
            )

        # Fuzzy name check (skip if user confirmed with force=true)
        if not force:
            existing_names = [
                (m.id, f"{m.first_name} {m.last_name}")
                for m in session.query(RosterMember).all()
            ]
            for mid, name in existing_names:
                ratio = difflib.SequenceMatcher(None, full_name, name.lower()).ratio()
                if ratio >= 0.8:
                    return RedirectResponse(
                        f"/settings?roster_fuzzy={name}&roster_fn={first_name}"
                        f"&roster_ln={last_name}&roster_em={email}#roster",
                        status_code=303
                    )

        session.add(RosterMember(first_name=first_name, last_name=last_name, email=email))
        session.commit()

    return RedirectResponse("/settings?roster_added=1&roster_skipped=0", status_code=303)


@router.post("/roster/upload", response_class=HTMLResponse)
async def upload_roster(request: Request, file: UploadFile = File(...)):
    """Upload an Excel file to populate the roster."""
    import io

    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        return RedirectResponse("/settings?roster_error=Please+upload+an+Excel+file+(.xlsx)", status_code=303)

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Find header row — look for first_name/last_name/email columns (case-insensitive)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        val = str(cell.value or "").strip().lower().replace(" ", "_")
        if val in ("first_name", "firstname", "first"):
            headers["first_name"] = col_idx
        elif val in ("last_name", "lastname", "last"):
            headers["last_name"] = col_idx
        elif val in ("email", "email_address", "work_email"):
            headers["email"] = col_idx

    if not all(k in headers for k in ("first_name", "last_name", "email")):
        return RedirectResponse(
            "/settings?roster_error=Excel+must+have+columns:+first_name,+last_name,+email",
            status_code=303
        )

    added = 0
    skipped = 0
    with get_session() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            first = str(row[headers["first_name"] - 1] or "").strip()
            last = str(row[headers["last_name"] - 1] or "").strip()
            email = str(row[headers["email"] - 1] or "").strip().lower()

            if not first or not last or not email or "@" not in email:
                skipped += 1
                continue

            existing = session.query(RosterMember).filter_by(email=email).first()
            if existing:
                existing.first_name = first
                existing.last_name = last
                existing.active = True
                skipped += 1
            else:
                session.add(RosterMember(first_name=first, last_name=last, email=email))
                added += 1
        session.commit()

    return RedirectResponse(f"/settings?roster_added={added}&roster_skipped={skipped}", status_code=303)


@router.post("/roster/{member_id}/delete")
async def delete_roster_member(member_id: int):
    """Remove a member from the roster."""
    with get_session() as session:
        member = session.query(RosterMember).filter_by(id=member_id).first()
        if member:
            name = member.full_name
            session.delete(member)
            session.commit()
            return RedirectResponse(f"/settings?roster_deleted={name}#roster", status_code=303)
    return RedirectResponse("/settings#roster", status_code=303)
