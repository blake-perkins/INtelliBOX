"""FastAPI web application for INtelliBOX."""

import threading
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import case, desc, func, or_
from sqlalchemy.orm import joinedload

from intellibox.database import get_session
from intellibox.knowledge import search_knowledge_base
from intellibox.models import Action, Assignment, Email, KnowledgeChunk, KnowledgeDocument, RosterMember
from intellibox.reporter.generator import (
    generate_enhanced_report,
    get_cached_structured_program_news,
)
from intellibox.settings_service import SettingsService
from intellibox.utils.datetime_utils import utcnow


def _start_background_watcher():
    """Start the file watcher in a background daemon thread with supervised restarts."""
    from intellibox.ai.processor import process_email_with_ai
    from intellibox.ingestion.file_watcher import supervised_watch
    from intellibox.ingestion.parser import parse_and_store_email

    inbox_dir = Path("data/inbox")

    def callback(eml_path: Path):
        with get_session() as session:
            email_record = parse_and_store_email(eml_path, session)
            if email_record:
                process_email_with_ai(email_record, session)

    thread = threading.Thread(
        target=supervised_watch,
        args=(inbox_dir, callback),
        kwargs={"interval": 5, "max_restarts": 5},
        daemon=True,
        name="inbox-watcher",
    )
    thread.start()


@asynccontextmanager
async def lifespan(app: FastAPI):
    _start_background_watcher()
    yield


# Create FastAPI app
app = FastAPI(
    title="INtelliBOX",
    description="AI-Powered Email Action Tracking System",
    version="1.0.0",
    lifespan=lifespan,
)

# Setup templates and static files
template_dir = Path(__file__).parent / "templates"
static_dir = Path(__file__).parent / "static"
templates = Jinja2Templates(directory=str(template_dir))
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

# Inject as get_program_name to avoid clashing with context vars named program_name
templates.env.globals['get_program_name'] = lambda: SettingsService.get_setting('program_name', '')


@app.middleware("http")
async def no_cache_html(request: Request, call_next):
    """Prevent browser caching of HTML pages so data is always fresh."""
    response = await call_next(request)
    ct = response.headers.get("content-type", "")
    if "text/html" in ct:
        response.headers["Cache-Control"] = "no-store"
    return response


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    """Main dashboard showing summary statistics."""
    with get_session() as session:
        # Get statistics — single aggregated query
        total_emails = session.query(func.count(Email.id)).scalar()
        stats = session.query(
            func.count(Action.id).label("total_actions"),
            func.count(case((Assignment.id.is_(None), Action.id))).label("unassigned"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "high"), Action.id))).label("high"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "medium"), Action.id))).label("medium"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "low"), Action.id))).label("low"),
            func.count(case((Assignment.id.isnot(None), Action.id))).label("assigned"),
            func.count(case((Assignment.status == "completed", Action.id))).label("completed"),
        ).select_from(Action).outerjoin(Assignment).one()

        total_actions = stats.total_actions
        unassigned_actions = stats.unassigned
        high_priority = stats.high
        medium_priority = stats.medium
        low_priority = stats.low
        assigned_actions = stats.assigned
        completed_actions = stats.completed

        # Get overdue actions (assigned or unassigned, exclude completed)
        today = utcnow().date()
        overdue_count = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()
        overdue_actions = session.query(Action).options(
            joinedload(Action.assignments), joinedload(Action.email),
        ).outerjoin(Assignment).join(Email).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).order_by(Action.due_date).all()

        # Get high priority unassigned actions
        high_priority_actions = session.query(Action).options(
            joinedload(Action.assignments), joinedload(Action.email),
        ).outerjoin(Assignment).join(Email).filter(
            Assignment.id.is_(None),
            Action.priority == "high"
        ).order_by(Action.due_date.asc().nullslast()).all()

        # Get medium and low priority unassigned actions
        medium_priority_actions = session.query(Action).options(
            joinedload(Action.assignments), joinedload(Action.email),
        ).outerjoin(Assignment).join(Email).filter(
            Assignment.id.is_(None),
            Action.priority == "medium"
        ).order_by(Action.due_date.asc().nullslast()).all()

        low_priority_actions = session.query(Action).options(
            joinedload(Action.assignments), joinedload(Action.email),
        ).outerjoin(Assignment).join(Email).filter(
            Assignment.id.is_(None),
            Action.priority == "low"
        ).order_by(Action.due_date.asc().nullslast()).all()

        # Get recent assignments (last 5, exclude completed)
        recent_assignments = session.query(Assignment, Action).join(
            Action
        ).filter(
            Assignment.status == "assigned"
        ).order_by(desc(Assignment.assigned_at)).limit(5).all()

        # Get recently completed (last 5)
        recent_completions = session.query(Assignment, Action).join(
            Action
        ).filter(
            Assignment.status == "completed"
        ).order_by(desc(Assignment.assigned_at)).limit(5).all()

        # Get roster for quick-assign dropdowns
        roster = session.query(RosterMember).filter_by(active=True).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()

        # Get latest email date for "last updated" info
        latest_email = session.query(Email).order_by(desc(Email.received_date)).first()
        last_email_date = latest_email.received_date if latest_email else None

        # Calculate time since last email
        time_since_last_email = None
        if last_email_date:
            delta = utcnow() - last_email_date
            if delta.total_seconds() < 3600:
                time_since_last_email = f"{int(delta.total_seconds() / 60)}m ago"
            elif delta.total_seconds() < 86400:
                time_since_last_email = f"{int(delta.total_seconds() / 3600)}h ago"
            else:
                time_since_last_email = f"{int(delta.total_seconds() / 86400)}d ago"

        return templates.TemplateResponse("dashboard.html", {
            "request": request,
            "total_emails": total_emails,
            "total_actions": total_actions,
            "unassigned_actions": unassigned_actions,
            "assigned_actions": assigned_actions,
            "completed_actions": completed_actions,
            "high_priority": high_priority,
            "medium_priority": medium_priority,
            "low_priority": low_priority,
            "overdue_count": overdue_count,
            "overdue_actions": overdue_actions,
            "high_priority_actions": high_priority_actions,
            "medium_priority_actions": medium_priority_actions,
            "low_priority_actions": low_priority_actions,
            "roster": roster,
            "recent_assignments": recent_assignments,
            "recent_completions": recent_completions,
            "last_email_date": last_email_date,
            "time_since_last_email": time_since_last_email,
            "current_time": utcnow(),
        })


@app.get("/actions", response_class=HTMLResponse)
async def list_actions(
    request: Request,
    priority: Optional[str] = None,
    assigned: Optional[str] = None,
    completed: Optional[str] = None,
    assignee: Optional[str] = None,
    search: Optional[str] = None,
    overdue: Optional[str] = None,
    page: int = Query(1, ge=1)
):
    """List all actions with filtering."""
    with get_session() as session:
        # Get stats for header — single aggregated query
        today = utcnow().date()
        action_stats = session.query(
            func.count(Action.id).label("total"),
            func.count(case((Assignment.id.is_(None), Action.id))).label("unassigned"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "high"), Action.id))).label("high"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "medium"), Action.id))).label("medium"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "low"), Action.id))).label("low"),
            func.count(case((
                (Action.due_date < today) & ((Assignment.id.is_(None)) | (Assignment.status != "completed")), Action.id
            ))).label("overdue"),
        ).select_from(Action).outerjoin(Assignment).one()

        total_actions = action_stats.total
        unassigned_actions = action_stats.unassigned
        high_priority = action_stats.high
        medium_priority = action_stats.medium
        low_priority = action_stats.low
        overdue_count = action_stats.overdue

        # Base query with eager loading to prevent N+1 in templates
        query = session.query(Action).options(
            joinedload(Action.assignments),
            joinedload(Action.email),
        ).outerjoin(Assignment).join(Email)

        # Apply filters
        if priority:
            query = query.filter(Action.priority == priority)

        if completed == "true" or assigned == "completed":
            query = query.filter(Assignment.status == "completed")
        elif assigned == "true":
            query = query.filter(Assignment.id.isnot(None))
        elif assigned == "false":
            query = query.filter(Assignment.id.is_(None))

        if assignee:
            query = query.filter(Assignment.assigned_to == assignee)

        if overdue == "true":
            query = query.filter(Action.due_date < today)

        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    Action.title.ilike(search_term),
                    Action.description.ilike(search_term),
                    Email.subject.ilike(search_term)
                )
            )

        # Order by priority and due date
        priority_order = case(
            (Action.priority == "high", 1),
            (Action.priority == "medium", 2),
            (Action.priority == "low", 3),
            else_=4
        )
        query = query.order_by(priority_order, Action.due_date.asc().nullslast())

        # Pagination
        per_page = 50
        total_count = query.count()
        total_pages = (total_count + per_page - 1) // per_page

        actions = query.offset((page - 1) * per_page).limit(per_page).all()

        # Get list of assignees for filter dropdown
        assignees = session.query(Assignment.assigned_to).distinct().order_by(Assignment.assigned_to).all()
        assignee_list = [a[0] for a in assignees if a[0]]

        # Get roster members for Quick Assign dropdown
        roster = session.query(RosterMember).filter_by(active=True).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()

        return templates.TemplateResponse("actions.html", {
            "request": request,
            "actions": actions,
            "priority": priority,
            "assigned": assigned,
            "completed": completed,
            "assignee": assignee,
            "search": search,
            "page": page,
            "total_pages": total_pages,
            "total_count": total_count,
            "total_actions": total_actions,
            "unassigned_actions": unassigned_actions,
            "high_priority": high_priority,
            "medium_priority": medium_priority,
            "low_priority": low_priority,
            "overdue_count": overdue_count,
            "overdue": overdue,
            "assignee_list": assignee_list,
            "roster": roster,
            "current_time": utcnow()
        })


@app.get("/actions/{action_id}", response_class=HTMLResponse)
async def view_action(request: Request, action_id: int):
    """View details of a specific action."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()

        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        # Get assignment if exists
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()

        # Get other actions from the same email
        related_actions = session.query(Action).filter(
            Action.email_id == action.email_id,
            Action.id != action_id
        ).all()

        # Get roster for assignment dropdown
        roster = session.query(RosterMember).filter_by(active=True).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()

        # Get recent assignees as fallback when no roster
        recent_assignees = session.query(Assignment.assigned_to).distinct().order_by(
            desc(Assignment.assigned_at)
        ).limit(10).all()
        recent_assignee_list = [a[0] for a in recent_assignees if a[0]]

        ai_config = SettingsService.get_ai_config()
        categories = ai_config.get('categories', SettingsService.DEFAULT_CATEGORIES)

        # Search KB for program context related to this action
        kb_matches = search_knowledge_base(
            sender=action.email.from_address,
            title=action.title,
            description=action.description or "",
            category=action.category or "",
        )

        return templates.TemplateResponse("action_detail.html", {
            "request": request,
            "action": action,
            "assignment": assignment,
            "related_actions": related_actions,
            "roster": roster,
            "recent_assignees": recent_assignee_list,
            "current_time": utcnow(),
            "categories": categories,
            "kb_matches": kb_matches,
        })


@app.get("/emails", response_class=HTMLResponse)
async def list_emails(
    request: Request,
    search: Optional[str] = None,
    processed: Optional[str] = None,
    days: Optional[str] = None,
    page: int = Query(1, ge=1)
):
    """List all emails."""
    days = int(days) if days and days.strip().isdigit() else None
    with get_session() as session:
        # Shared banner stats (same as all pages)
        unassigned_actions = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None)
        ).count()
        high_priority = session.query(Action).outerjoin(Assignment).filter(
            Action.priority == "high",
            Assignment.id.is_(None)
        ).count()
        today = utcnow().date()
        overdue_count = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()
        latest_email = session.query(Email).order_by(desc(Email.received_date)).first()
        last_email_date = latest_email.received_date if latest_email else None
        time_since_last_email = None
        if last_email_date:
            delta = utcnow() - last_email_date
            if delta.total_seconds() < 3600:
                time_since_last_email = f"{int(delta.total_seconds() / 60)}m ago"
            elif delta.total_seconds() < 86400:
                time_since_last_email = f"{int(delta.total_seconds() / 3600)}h ago"
            else:
                time_since_last_email = f"{int(delta.total_seconds() / 86400)}d ago"

        # Get stats
        total_emails = session.query(Email).count()
        processed_emails = session.query(Email).filter(Email.processed == True).count()
        unprocessed_emails = total_emails - processed_emails

        # Count emails with high priority actions
        emails_with_high_priority = session.query(Email).join(Action).filter(
            Action.priority == "high"
        ).distinct().count()

        # Base query
        query = session.query(Email)

        # Apply filters
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    Email.subject.ilike(search_term),
                    Email.from_address.ilike(search_term),
                    Email.body_text.ilike(search_term)
                )
            )

        if processed == "true":
            query = query.filter(Email.processed == True)
        elif processed == "false":
            query = query.filter(Email.processed == False)

        # Date range filter
        if days:
            cutoff_date = utcnow() - timedelta(days=days)
            query = query.filter(Email.received_date >= cutoff_date)

        # Order by received date
        query = query.order_by(desc(Email.received_date))

        # Pagination
        per_page = 50
        total_count = query.count()
        total_pages = (total_count + per_page - 1) // per_page

        emails = query.offset((page - 1) * per_page).limit(per_page).all()

        return templates.TemplateResponse("emails.html", {
            "request": request,
            "emails": emails,
            "search": search,
            "processed": processed,
            "days": days,
            "page": page,
            "total_pages": total_pages,
            "total_count": total_count,
            "total_emails": total_emails,
            "processed_emails": processed_emails,
            "unprocessed_emails": unprocessed_emails,
            "emails_with_high_priority": emails_with_high_priority,
            "unassigned_actions": unassigned_actions,
            "high_priority": high_priority,
            "overdue_count": overdue_count,
            "time_since_last_email": time_since_last_email,
            "current_time": utcnow()
        })


ALLOWED_EMAIL_EXTENSIONS = {".eml", ".msg"}
MAX_EMAIL_FILE_SIZE = 25 * 1024 * 1024  # 25 MB


@app.post("/emails/upload")
async def upload_email(file: UploadFile = File(...)):
    """Upload an .eml or .msg file for processing."""
    filename = file.filename or "unknown"
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ALLOWED_EMAIL_EXTENSIONS:
        return RedirectResponse(
            "/emails?upload_error=Unsupported+file+type.+Please+upload+.eml+or+.msg",
            status_code=303,
        )

    content = await file.read()
    if len(content) > MAX_EMAIL_FILE_SIZE:
        return RedirectResponse(
            "/emails?upload_error=File+too+large.+Maximum+size+is+25+MB",
            status_code=303,
        )
    if len(content) == 0:
        return RedirectResponse(
            "/emails?upload_error=File+is+empty",
            status_code=303,
        )

    inbox_dir = Path("data/inbox")
    inbox_dir.mkdir(parents=True, exist_ok=True)
    dest = inbox_dir / filename

    # Avoid overwriting existing files
    if dest.exists():
        stem = dest.stem
        suffix = dest.suffix
        counter = 1
        while dest.exists():
            dest = inbox_dir / f"{stem}_{counter}{suffix}"
            counter += 1

    dest.write_bytes(content)
    return RedirectResponse("/emails?uploaded=1", status_code=303)


@app.get("/emails/{email_id}", response_class=HTMLResponse)
async def view_email(request: Request, email_id: int):
    """View details of a specific email."""
    with get_session() as session:
        email = session.query(Email).filter_by(id=email_id).first()

        if not email:
            raise HTTPException(status_code=404, detail="Email not found")

        # Get associated actions
        actions = session.query(Action).filter_by(email_id=email_id).all()

        # Count actions by status
        assigned_count = sum(1 for action in actions if action.assignments)
        unassigned_count = len(actions) - assigned_count
        completed_count = sum(1 for action in actions if action.assignments and action.assignments[0].status == 'completed')

        # Count by priority
        high_priority_count = sum(1 for action in actions if action.priority == 'high')

        # Parse also_received_from JSON for template
        import json as _json
        also_received = _json.loads(email.also_received_from) if email.also_received_from else []

        return templates.TemplateResponse("email_detail.html", {
            "request": request,
            "email": email,
            "actions": actions,
            "assigned_count": assigned_count,
            "unassigned_count": unassigned_count,
            "completed_count": completed_count,
            "high_priority_count": high_priority_count,
            "also_received": also_received,
            "current_time": utcnow()
        })


@app.get("/insights", response_class=HTMLResponse)
async def view_insights(request: Request):
    """View AI-powered insights dashboard. Always loads instantly using cached data."""
    try:
        days = int(request.query_params.get("days", 14))
    except (ValueError, TypeError):
        days = 14
    days = max(7, min(days, 90))  # clamp to 7-90

    with get_session() as session:
        # Always use cached data — never block page load with AI calls
        report_data = generate_enhanced_report(session, days=days, force_refresh=False)

        # Calculate cache age in minutes for display
        if report_data.get("is_cached") and report_data.get("generated_at"):
            cache_age_minutes = int((utcnow() - report_data["generated_at"]).total_seconds() / 60)
        else:
            cache_age_minutes = 0

        program_news_data = get_cached_structured_program_news(session)

        return templates.TemplateResponse("report.html", {
            "request": request,
            "report": report_data,
            "cache_age_minutes": cache_age_minutes,
            "program_news": program_news_data,
            "current_time": utcnow(),
            "datetime": datetime,
            "selected_days": days
        })


@app.post("/api/insights-refresh")
async def refresh_insights(request: Request):
    """Regenerate AI insights and program news. Called via AJAX from the Insights page."""
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    days = int(body.get("days", 14))
    days = max(7, min(days, 90))

    with get_session() as session:
        generate_enhanced_report(session, days=days, force_refresh=True)
        get_cached_structured_program_news(session, days=days, force_refresh=True)
    return JSONResponse({"status": "ok"})


@app.get("/api/stats")
async def get_stats():
    """API endpoint for statistics (for auto-refresh)."""
    with get_session() as session:
        today = utcnow().date()
        last_sync = "never"
        last_sync_str = SettingsService.get_setting('last_sync_time', None)
        if last_sync_str:
            try:
                last_sync_dt = datetime.fromisoformat(last_sync_str)
                delta = utcnow() - last_sync_dt
                if delta.total_seconds() < 3600:
                    last_sync = f"{int(delta.total_seconds() / 60)}m ago"
                elif delta.total_seconds() < 86400:
                    last_sync = f"{int(delta.total_seconds() / 3600)}h ago"
                else:
                    last_sync = f"{int(delta.total_seconds() / 86400)}d ago"
            except (ValueError, TypeError):
                pass
        # Single aggregated query for all action/assignment stats
        s = session.query(
            func.count(Action.id).label("total_actions"),
            func.count(case((Assignment.id.is_(None), Action.id))).label("unassigned"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "high"), Action.id))).label("unassigned_high"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "medium"), Action.id))).label("unassigned_medium"),
            func.count(case(((Assignment.id.is_(None)) & (Action.priority == "low"), Action.id))).label("unassigned_low"),
            func.count(case((
                (Action.due_date < today) & ((Assignment.id.is_(None)) | (Assignment.status != "completed")), Action.id
            ))).label("overdue"),
            func.count(case(((Assignment.id.isnot(None)) & (Assignment.status != "completed"), Action.id))).label("assigned"),
            func.count(case((
                (Assignment.status != "completed") & (Action.priority == "high") & (Assignment.id.isnot(None)), Action.id
            ))).label("assigned_high"),
            func.count(case((
                (Assignment.status != "completed") & (Action.priority == "medium") & (Assignment.id.isnot(None)), Action.id
            ))).label("assigned_medium"),
            func.count(case((
                (Assignment.status != "completed") & (Action.priority == "low") & (Assignment.id.isnot(None)), Action.id
            ))).label("assigned_low"),
            func.count(case((Assignment.status == "completed", Action.id))).label("completed"),
        ).select_from(Action).outerjoin(Assignment).one()

        total_emails = session.query(func.count(Email.id)).scalar()

        return {
            "last_sync": last_sync,
            "total_emails": total_emails,
            "total_actions": s.total_actions,
            "unassigned_actions": s.unassigned,
            "unassigned_high": s.unassigned_high,
            "unassigned_medium": s.unassigned_medium,
            "unassigned_low": s.unassigned_low,
            "overdue_count": s.overdue,
            "assigned_actions": s.assigned,
            "assigned_high": s.assigned_high,
            "assigned_medium": s.assigned_medium,
            "assigned_low": s.assigned_low,
            "completed_actions": s.completed,
        }


@app.post("/actions/{action_id}/assign")
async def assign_action(
    action_id: int,
    assigned_to: str = Form(...),
    notes: str = Form(""),
):
    """Assign an action to someone."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        # Check if already assigned
        existing = session.query(Assignment).filter_by(action_id=action_id).first()

        if existing:
            # Update existing assignment
            existing.assigned_to = assigned_to
            existing.notes = notes
            existing.assigned_at = utcnow()
        else:
            # Create new assignment
            assignment = Assignment(
                action_id=action_id,
                assigned_to=assigned_to,
                notes=notes,
                status="assigned"
            )
            session.add(assignment)

        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/complete")
async def complete_action(action_id: int):
    """Mark an action as completed."""
    with get_session() as session:
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()

        if not assignment:
            # Create assignment with completed status
            assignment = Assignment(
                action_id=action_id,
                assigned_to="Unknown",
                status="completed",
                completed_at=utcnow()
            )
            session.add(assignment)
        else:
            assignment.status = "completed"
            assignment.completed_at = utcnow()

        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)



@app.post("/actions/{action_id}/priority")
async def change_priority(action_id: int, priority: str = Form(...)):
    """Quick-change priority (used by dashboard AJAX)."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")
        if priority in ["high", "medium", "low"]:
            action.priority = priority
            session.commit()
    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/unassign")
async def unassign_action(action_id: int):
    """Remove assignment from an action (used by dashboard AJAX)."""
    with get_session() as session:
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()
        if assignment:
            session.delete(assignment)
            session.commit()
    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/status")
async def update_assignment_status(
    action_id: int,
    status: str = Form(...)
):
    """Update assignment status (assigned, completed)."""
    with get_session() as session:
        assignment = session.query(Assignment).filter_by(action_id=action_id).first()

        if not assignment:
            raise HTTPException(status_code=404, detail="Action not assigned")

        if status not in ["assigned", "in_progress", "completed"]:
            raise HTTPException(status_code=400, detail="Invalid status")

        assignment.status = status
        if status == "completed":
            assignment.completed_at = utcnow()
        else:
            assignment.completed_at = None
        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/edit")
async def edit_action(
    action_id: int,
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form("medium"),
    due_date: str = Form(""),
    category: str = Form(""),
    assigned_to: str = Form(""),
    notes: str = Form(""),
):
    """Update action fields and optionally assign in one submit."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        # Block edits on completed actions — must reopen first
        existing_assignment = session.query(Assignment).filter_by(action_id=action_id).first()
        if existing_assignment and existing_assignment.status == "completed":
            return RedirectResponse(url=f"/actions/{action_id}", status_code=303)

        action.title = title.strip()
        action.description = description.strip() if description.strip() else None
        action.category = category.strip() if category.strip() else None

        if priority in ["high", "medium", "low"]:
            action.priority = priority

        if due_date and due_date.strip():
            from datetime import datetime as dt
            try:
                action.due_date = dt.strptime(due_date.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass
        else:
            action.due_date = None

        # Handle assignment
        existing = session.query(Assignment).filter_by(action_id=action_id).first()
        if assigned_to and assigned_to.strip():
            if existing:
                existing.assigned_to = assigned_to.strip()
                existing.notes = notes.strip() if notes.strip() else existing.notes
                existing.assigned_at = utcnow()
            else:
                new_assignment = Assignment(
                    action_id=action_id,
                    assigned_to=assigned_to.strip(),
                    notes=notes.strip() if notes.strip() else None,
                    status="assigned",
                )
                session.add(new_assignment)
        elif existing:
            # User selected "Unassigned" — remove the assignment
            session.delete(existing)

        session.commit()

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.post("/actions/{action_id}/delete")
async def delete_action(action_id: int):
    """Delete an action."""
    with get_session() as session:
        action = session.query(Action).filter_by(id=action_id).first()
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")

        email_id = action.email_id
        session.delete(action)
        session.commit()

    return RedirectResponse(url=f"/emails/{email_id}", status_code=303)


@app.get("/emails/{email_id}/actions/new", response_class=HTMLResponse)
async def new_action_form(request: Request, email_id: int):
    """Show form to create a new action for an email."""
    with get_session() as session:
        email = session.query(Email).filter_by(id=email_id).first()
        if not email:
            raise HTTPException(status_code=404, detail="Email not found")

        ai_config = SettingsService.get_ai_config()
        categories = ai_config.get('categories', SettingsService.DEFAULT_CATEGORIES)
        return templates.TemplateResponse("action_new.html", {
            "request": request,
            "email": email,
            "categories": categories,
        })


@app.post("/emails/{email_id}/actions/new")
async def create_action(
    email_id: int,
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form("medium"),
    due_date: Optional[str] = Form(None),
    category: str = Form("")
):
    """Create a new action for an email."""
    with get_session() as session:
        email = session.query(Email).filter_by(id=email_id).first()
        if not email:
            raise HTTPException(status_code=404, detail="Email not found")

        # Parse due date if provided
        parsed_due_date = None
        if due_date and due_date.strip():
            try:
                from datetime import datetime as dt
                parsed_due_date = dt.strptime(due_date, '%Y-%m-%d')
            except ValueError:
                pass

        # Create new action
        action = Action(
            email_id=email_id,
            title=title,
            description=description if description.strip() else None,
            priority=priority,
            due_date=parsed_due_date,
            category=category if category.strip() else None,
            confidence_score=1.0  # Manual actions are 100% confident
        )

        session.add(action)
        session.commit()
        session.refresh(action)
        action_id = action.id  # capture before session closes

    return RedirectResponse(url=f"/actions/{action_id}", status_code=303)


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, success: bool = False):
    """Settings page for configuring priority rules."""
    priority_config = SettingsService.get_priority_config()
    ai_config = SettingsService.get_ai_config()
    high_senders = priority_config.get('high_senders', [])
    high_keywords = priority_config.get('high_keywords', [])
    categories = ai_config.get('categories', SettingsService.DEFAULT_CATEGORIES)

    with get_session() as session:
        roster = session.query(RosterMember).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()

        # Knowledge Base data
        kb_documents = session.query(KnowledgeDocument).order_by(
            desc(KnowledgeDocument.uploaded_at)
        ).all()
        kb_total_size_bytes = sum(d.file_size for d in kb_documents)
        kb_total_chars = sum(d.text_length for d in kb_documents)
        if kb_total_size_bytes < 1024:
            kb_total_size = f"{kb_total_size_bytes} B"
        elif kb_total_size_bytes < 1024 * 1024:
            kb_total_size = f"{kb_total_size_bytes / 1024:.1f} KB"
        else:
            kb_total_size = f"{kb_total_size_bytes / (1024 * 1024):.1f} MB"

        return templates.TemplateResponse(
            "settings.html",
            {
                "request": request,
                "priority_default": priority_config.get('default_priority', 'medium'),
                "priority_days_threshold": priority_config.get('days_threshold', 5),
                "priority_high_senders_text": '\n'.join(high_senders),
                "priority_high_keywords_text": '\n'.join(high_keywords),
                "confidence_threshold": ai_config.get('confidence_threshold', 0.5),
                "categories": categories,
                "timezone": SettingsService.get_timezone(),
                "program_name": SettingsService.get_setting('program_name', ''),
                "insights_prompt": SettingsService.get_insights_prompt(),
                "success": success,
                "roster": roster,
                "kb_documents": kb_documents,
                "kb_total_size": kb_total_size,
                "kb_total_chars": kb_total_chars,
            }
        )


@app.post("/settings")
async def save_settings(
    request: Request,
    priority_default: str = Form(...),
    priority_days_threshold: int = Form(...),
    priority_high_senders: str = Form(""),
    priority_high_keywords: str = Form(""),
    confidence_threshold: float = Form(0.5),
    timezone: str = Form("America/Chicago"),
    program_name: str = Form("")
):
    """Save priority and AI settings."""
    # Parse textarea inputs (newline-separated) into lists
    senders = [s.strip() for s in priority_high_senders.split('\n') if s.strip()]
    keywords = [k.strip() for k in priority_high_keywords.split('\n') if k.strip()]

    # Save settings
    SettingsService.set_setting('priority_default', priority_default)
    SettingsService.set_setting('priority_days_threshold', priority_days_threshold)
    SettingsService.set_setting('priority_high_senders', senders)
    SettingsService.set_setting('priority_high_keywords', keywords)
    SettingsService.set_setting('ai_confidence_threshold', round(float(confidence_threshold), 2))
    SettingsService.set_setting('timezone', timezone)
    SettingsService.set_setting('program_name', program_name.strip())

    # Redirect with success flag
    return RedirectResponse(url="/settings?success=true", status_code=303)


@app.post("/settings/insights-prompt")
async def save_insights_prompt(insights_prompt: str = Form(...)):
    """Save custom insights prompt template."""
    SettingsService.set_setting(
        "insights_prompt", insights_prompt,
        description="Custom prompt template for Insights page AI analysis"
    )
    return RedirectResponse(url="/settings?tab=prompt&prompt_saved=1", status_code=303)


@app.post("/settings/insights-prompt/reset")
async def reset_insights_prompt():
    """Reset insights prompt to hardcoded default."""
    SettingsService.delete_setting("insights_prompt")
    return JSONResponse({"status": "ok"})


@app.post("/categories/add")
async def add_category(name: str = Form(...), description: str = Form("")):
    """Add a new action category."""
    name = name.strip()
    if not name:
        return RedirectResponse(url="/settings?tab=categories&cat_error=Name+is+required", status_code=303)
    ai_config = SettingsService.get_ai_config()
    categories = ai_config.get('categories', list(SettingsService.DEFAULT_CATEGORIES))
    if any(c['name'].lower() == name.lower() for c in categories):
        return RedirectResponse(url=f"/settings?tab=categories&cat_error=Category+%27{name}%27+already+exists", status_code=303)
    categories.append({"name": name, "description": description.strip()})
    SettingsService.set_setting('ai_categories', categories)
    return RedirectResponse(url=f"/settings?tab=categories&cat_added={name}", status_code=303)


@app.post("/categories/delete")
async def delete_category(name: str = Form(...)):
    """Remove an action category."""
    ai_config = SettingsService.get_ai_config()
    categories = ai_config.get('categories', list(SettingsService.DEFAULT_CATEGORIES))
    categories = [c for c in categories if c['name'] != name]
    SettingsService.set_setting('ai_categories', categories)
    return RedirectResponse(url=f"/settings?tab=categories&cat_deleted={name}", status_code=303)


@app.get("/analytics", response_class=HTMLResponse)
async def view_analytics(request: Request):
    """System analytics dashboard — all-time stats and activity trends."""
    with get_session() as session:
        now = utcnow()
        today = now.date()

        # --- All-time counters ---
        total_emails = session.query(Email).count()
        total_actions = session.query(Action).count()
        total_completed = session.query(Assignment).filter(
            Assignment.status == "completed"
        ).count()
        total_members = session.query(RosterMember).count()

        # --- Current pipeline ---
        unassigned = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None)
        ).count()
        in_progress = session.query(Assignment).filter(
            Assignment.status != "completed"
        ).count()
        overdue = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()

        # --- Weekly activity (last 4 weeks) ---
        weeks = []
        for i in range(4):
            week_end = now - timedelta(days=i * 7)
            week_start = week_end - timedelta(days=7)
            emails_count = session.query(Email).filter(
                Email.created_at >= week_start,
                Email.created_at < week_end
            ).count()
            actions_count = session.query(Action).filter(
                Action.created_at >= week_start,
                Action.created_at < week_end
            ).count()
            completed_count = session.query(Assignment).filter(
                Assignment.status == "completed",
                Assignment.completed_at >= week_start,
                Assignment.completed_at < week_end
            ).count()
            weeks.append({
                "label": week_start.strftime("%b %d") + " – " + (week_end - timedelta(days=1)).strftime("%b %d"),
                "emails": emails_count,
                "actions": actions_count,
                "completed": completed_count,
            })
        weeks.reverse()  # oldest first

        # --- Top categories (top 5) ---
        category_rows = session.query(
            Action.category, func.count(Action.id).label("cnt")
        ).filter(Action.category.isnot(None), Action.category != "").group_by(
            Action.category
        ).order_by(desc("cnt")).limit(5).all()
        top_categories = [{"name": r[0], "count": r[1]} for r in category_rows]

        # --- Team activity ---
        team_rows = session.query(
            Assignment.assigned_to,
            func.count(case((Assignment.status == "completed", 1))).label("done"),
            func.count(case((Assignment.status != "completed", 1))).label("active"),
        ).group_by(Assignment.assigned_to).order_by(desc("done")).all()
        team_activity = [{"name": r[0], "completed": r[1], "active": r[2]} for r in team_rows]

        # Max values for bar scaling
        max_weekly = max(
            (max(w["emails"], w["actions"], w["completed"]) for w in weeks),
            default=0
        )

        return templates.TemplateResponse("analytics.html", {
            "request": request,
            "total_emails": total_emails,
            "total_actions": total_actions,
            "total_completed": total_completed,
            "total_members": total_members,
            "unassigned": unassigned,
            "in_progress": in_progress,
            "completed_all": total_completed,
            "overdue": overdue,
            "weeks": weeks,
            "max_weekly": max_weekly,
            "top_categories": top_categories,
            "team_activity": team_activity,
        })


# --- Knowledge Base ---

ALLOWED_KB_EXTENSIONS = {".pdf": "pdf", ".docx": "docx", ".txt": "txt"}
MAX_KB_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


@app.get("/knowledge-base", response_class=HTMLResponse)
async def knowledge_base(request: Request):
    """Redirect legacy KB page to the Settings Knowledge Base tab."""
    return RedirectResponse("/settings?tab=kb", status_code=302)


@app.post("/knowledge-base/upload")
async def upload_knowledge_doc(
    file: UploadFile = File(...),
    description: str = Form(""),
):
    """Upload a document to the knowledge base."""
    from intellibox.knowledge.extractor import extract_text

    filename = file.filename or "unknown"
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ALLOWED_KB_EXTENSIONS:
        return RedirectResponse(
            "/settings?tab=kb&kb_error=Unsupported+file+type.+Please+upload+PDF,+DOCX,+or+TXT",
            status_code=303,
        )

    file_type = ALLOWED_KB_EXTENSIONS[ext]
    content = await file.read()
    file_size = len(content)

    if file_size > MAX_KB_FILE_SIZE:
        return RedirectResponse(
            "/settings?tab=kb&kb_error=File+too+large.+Maximum+size+is+10+MB",
            status_code=303,
        )

    if file_size == 0:
        return RedirectResponse(
            "/settings?tab=kb&kb_error=File+is+empty",
            status_code=303,
        )

    extracted_text, status = extract_text(content, file_type)

    with get_session() as session:
        doc = KnowledgeDocument(
            filename=filename,
            file_type=file_type,
            file_size=file_size,
            description=description.strip() if description.strip() else None,
            extracted_text=extracted_text if extracted_text else None,
            extraction_status=status,
        )
        session.add(doc)
        session.commit()
        doc_id = doc.id

    if status == "failed":
        return RedirectResponse(
            "/settings?tab=kb&kb_warning=File+uploaded+but+text+extraction+failed",
            status_code=303,
        )
    elif status == "partial":
        return RedirectResponse(
            "/settings?tab=kb&kb_warning=File+uploaded+but+only+partial+text+could+be+extracted",
            status_code=303,
        )

    # Compute embeddings if API key is available (non-blocking best-effort)
    from intellibox.knowledge.embeddings import embed_document
    chunk_count = embed_document(doc_id)
    if chunk_count > 0:
        return RedirectResponse(
            f"/settings?tab=kb&kb_success=1&embedded={chunk_count}",
            status_code=303,
        )

    return RedirectResponse("/settings?tab=kb&kb_success=1", status_code=303)


@app.get("/knowledge-base/{doc_id}", response_class=HTMLResponse)
async def knowledge_base_detail(request: Request, doc_id: int, highlight: Optional[str] = None):
    """View a knowledge base document and its extracted text."""
    with get_session() as session:
        doc = session.query(KnowledgeDocument).filter_by(id=doc_id).first()
        if not doc:
            return RedirectResponse("/settings?tab=kb", status_code=303)
        return templates.TemplateResponse("knowledge_base_detail.html", {
            "request": request,
            "doc": doc,
            "highlight": highlight or "",
        })


@app.post("/knowledge-base/{doc_id}/delete")
async def delete_knowledge_doc(doc_id: int):
    """Delete a document and its embedding chunks from the knowledge base."""
    from intellibox.knowledge.embeddings import remove_document_chunks
    remove_document_chunks(doc_id)
    with get_session() as session:
        doc = session.query(KnowledgeDocument).filter_by(id=doc_id).first()
        if doc:
            session.delete(doc)
            session.commit()
    return RedirectResponse("/settings?tab=kb&kb_deleted=1", status_code=303)


@app.get("/roster", response_class=HTMLResponse)
async def view_roster(request: Request):
    """View and manage the program roster."""
    with get_session() as session:
        members = session.query(RosterMember).order_by(
            RosterMember.last_name, RosterMember.first_name
        ).all()
        unassigned_actions = session.query(Action).outerjoin(Assignment).filter(
            Assignment.id.is_(None)
        ).count()
        high_priority = session.query(Action).outerjoin(Assignment).filter(
            Action.priority == "high", Assignment.id.is_(None)
        ).count()
        today = utcnow().date()
        overdue_count = session.query(Action).outerjoin(Assignment).filter(
            Action.due_date < today,
            (Assignment.id.is_(None)) | (Assignment.status != "completed")
        ).count()
        return templates.TemplateResponse("roster.html", {
            "request": request,
            "members": members,
            "unassigned_actions": unassigned_actions,
            "high_priority": high_priority,
            "overdue_count": overdue_count,
            "current_time": utcnow(),
        })


@app.post("/roster/add")
async def add_roster_member(
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    force: str = Form(""),
):
    """Manually add a single roster member with duplicate and fuzzy-name checking."""
    import difflib

    first_name = first_name.strip()
    last_name = last_name.strip()
    email = email.strip().lower()

    if not email:
        return RedirectResponse(
            "/settings?roster_error=Email+is+required#roster",
            status_code=303
        )

    full_name = f"{first_name} {last_name}".lower()

    with get_session() as session:
        # Exact email duplicate
        if session.query(RosterMember).filter_by(email=email).first():
            return RedirectResponse(
                f"/settings?roster_error={email}+is+already+in+the+roster#roster",
                status_code=303
            )

        # Fuzzy name check (skip if user confirmed with force=true)
        if not force:
            existing_names = [
                (m.id, f"{m.first_name} {m.last_name}")
                for m in session.query(RosterMember).all()
            ]
            for mid, name in existing_names:
                ratio = difflib.SequenceMatcher(None, full_name, name.lower()).ratio()
                if ratio >= 0.8:
                    return RedirectResponse(
                        f"/settings?roster_fuzzy={name}&roster_fn={first_name}"
                        f"&roster_ln={last_name}&roster_em={email}#roster",
                        status_code=303
                    )

        session.add(RosterMember(first_name=first_name, last_name=last_name, email=email))
        session.commit()

    return RedirectResponse("/settings?roster_added=1&roster_skipped=0", status_code=303)


@app.post("/roster/upload", response_class=HTMLResponse)
async def upload_roster(request: Request, file: UploadFile = File(...)):
    """Upload an Excel file to populate the roster."""
    import io

    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        return RedirectResponse("/settings?roster_error=Please+upload+an+Excel+file+(.xlsx)", status_code=303)

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Find header row — look for first_name/last_name/email columns (case-insensitive)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        val = str(cell.value or "").strip().lower().replace(" ", "_")
        if val in ("first_name", "firstname", "first"):
            headers["first_name"] = col_idx
        elif val in ("last_name", "lastname", "last"):
            headers["last_name"] = col_idx
        elif val in ("email", "email_address", "work_email"):
            headers["email"] = col_idx

    if not all(k in headers for k in ("first_name", "last_name", "email")):
        return RedirectResponse(
            "/settings?roster_error=Excel+must+have+columns:+first_name,+last_name,+email",
            status_code=303
        )

    added = 0
    skipped = 0
    with get_session() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            first = str(row[headers["first_name"] - 1] or "").strip()
            last = str(row[headers["last_name"] - 1] or "").strip()
            email = str(row[headers["email"] - 1] or "").strip().lower()

            if not first or not last or not email or "@" not in email:
                skipped += 1
                continue

            existing = session.query(RosterMember).filter_by(email=email).first()
            if existing:
                existing.first_name = first
                existing.last_name = last
                existing.active = True
                skipped += 1
            else:
                session.add(RosterMember(first_name=first, last_name=last, email=email))
                added += 1
        session.commit()

    return RedirectResponse(f"/settings?roster_added={added}&roster_skipped={skipped}", status_code=303)


@app.post("/roster/{member_id}/delete")
async def delete_roster_member(member_id: int):
    """Remove a member from the roster."""
    with get_session() as session:
        member = session.query(RosterMember).filter_by(id=member_id).first()
        if member:
            name = member.full_name
            session.delete(member)
            session.commit()
            return RedirectResponse(f"/settings?roster_deleted={name}#roster", status_code=303)
    return RedirectResponse("/settings#roster", status_code=303)


@app.get("/health")
async def health_check():
    """Health check endpoint with watcher status."""
    from intellibox.ingestion.file_watcher import get_watcher_health

    return {
        "status": "healthy",
        "timestamp": utcnow().isoformat(),
        "watcher": get_watcher_health(),
    }


# --- Test-only endpoints (gated by TESTING env var) ---
import os as _os

if _os.environ.get("TESTING", "").lower() in ("true", "1", "yes"):

    @app.post("/api/test/reset")
    async def test_reset_database():
        """Drop and recreate all tables. Only available when TESTING=true."""
        from intellibox.database import engine
        from intellibox.models import Base
        Base.metadata.drop_all(bind=engine)
        Base.metadata.create_all(bind=engine)
        return {"status": "reset", "tables": list(Base.metadata.tables.keys())}

    @app.post("/api/test/seed")
    async def test_seed_entity(request: Request):
        """Create a test entity. Body: {"type": "email|action|assignment|roster_member", "data": {...}}"""
        payload = await request.json()
        entity_type = payload["type"]
        data = payload["data"]

        # Parse ISO date strings into datetime objects
        _date_fields = {
            "received_date", "processed_at", "due_date", "assigned_at", "completed_at", "created_at"
        }
        for key in list(data.keys()):
            if key in _date_fields and isinstance(data[key], str):
                data[key] = datetime.fromisoformat(data[key])

        model_map = {
            "email": Email,
            "action": Action,
            "assignment": Assignment,
            "roster_member": RosterMember,
            "knowledge_document": KnowledgeDocument,
            "knowledge_chunk": KnowledgeChunk,
        }
        Model = model_map[entity_type]

        with get_session() as session:
            entity = Model(**data)
            session.add(entity)
            session.flush()
            entity_id = entity.id
        return {"id": entity_id}

    @app.get("/api/test/query/action/{action_id}")
    async def test_query_action(action_id: int):
        """Query an action by ID for test assertions."""
        with get_session() as session:
            action = session.query(Action).filter_by(id=action_id).first()
            if not action:
                return {"found": False}
            return {
                "found": True,
                "id": action.id,
                "title": action.title,
                "priority": action.priority,
                "due_date": action.due_date.isoformat() if action.due_date else None,
                "category": action.category,
                "has_assignments": bool(action.assignments and len(action.assignments) > 0),
            }

    @app.get("/api/test/query/assignment")
    async def test_query_assignment(action_id: int):
        """Query an active assignment by action_id for test assertions."""
        with get_session() as session:
            assignment = session.query(Assignment).filter_by(action_id=action_id).first()
            if not assignment:
                return {"found": False}
            return {
                "found": True,
                "id": assignment.id,
                "action_id": assignment.action_id,
                "assigned_to": assignment.assigned_to,
                "status": assignment.status,
            }

    @app.get("/api/test/query/roster/count")
    async def test_query_roster_count():
        """Count roster members for test assertions."""
        with get_session() as session:
            count = session.query(RosterMember).count()
            return {"count": count}

    @app.get("/api/test/query/roster-member/{member_id}")
    async def test_query_roster_member(member_id: int):
        """Query a roster member by ID for test assertions."""
        with get_session() as session:
            member = session.query(RosterMember).filter_by(id=member_id).first()
            if not member:
                return {"found": False}
            return {"found": True, "id": member.id}

    @app.get("/api/test/query/knowledge/count")
    async def test_query_knowledge_count():
        """Count knowledge documents for test assertions."""
        with get_session() as session:
            count = session.query(KnowledgeDocument).count()
            return {"count": count}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
